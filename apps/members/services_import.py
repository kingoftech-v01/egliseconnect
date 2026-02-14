"""Member import service for CSV/Excel files with validation preview."""
import csv
import io
from datetime import datetime

from django.utils.translation import gettext_lazy as _

from apps.core.constants import Roles, FamilyStatus, Province


class MemberImportService:
    """Handles CSV/Excel import of members with field mapping and validation."""

    VALID_EXTENSIONS = ['.csv', '.xlsx']

    # Map of internal field names to validation functions
    FIELD_VALIDATORS = {
        'first_name': lambda v: bool(v and v.strip()),
        'last_name': lambda v: bool(v and v.strip()),
        'email': lambda v: not v or '@' in v,
        'phone': lambda v: True,
        'birth_date': lambda v: True,
        'address': lambda v: True,
        'city': lambda v: True,
        'province': lambda v: not v or v.upper() in dict(Province.CHOICES),
        'postal_code': lambda v: True,
        'role': lambda v: not v or v in dict(Roles.CHOICES),
        'family_status': lambda v: not v or v in dict(FamilyStatus.CHOICES),
    }

    @classmethod
    def parse_file(cls, uploaded_file):
        """
        Parse a CSV or Excel file and return headers + rows.

        Returns:
            tuple: (headers: list[str], rows: list[dict])
        """
        filename = uploaded_file.name.lower()

        if filename.endswith('.csv'):
            return cls._parse_csv(uploaded_file)
        elif filename.endswith('.xlsx'):
            return cls._parse_excel(uploaded_file)
        else:
            raise ValueError(_('Format de fichier non supporté. Utilisez CSV ou XLSX.'))

    @classmethod
    def _parse_csv(cls, uploaded_file):
        """Parse a CSV file."""
        content = uploaded_file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        headers = reader.fieldnames or []
        rows = list(reader)
        return headers, rows

    @classmethod
    def _parse_excel(cls, uploaded_file):
        """Parse an Excel file."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError(_('openpyxl requis pour importer des fichiers Excel.'))

        wb = load_workbook(uploaded_file, read_only=True)
        ws = wb.active

        rows_data = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(cell) if cell else f'col_{j}' for j, cell in enumerate(row)]
            else:
                row_dict = {}
                for j, cell in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = str(cell) if cell is not None else ''
                rows_data.append(row_dict)

        wb.close()
        return headers, rows_data

    @classmethod
    def validate_preview(cls, rows, field_mapping):
        """
        Validate mapped rows and return preview with errors.

        Args:
            rows: list of dicts from parsed file
            field_mapping: dict mapping csv_column -> member_field

        Returns:
            list of dicts: [{'row_num': int, 'data': dict, 'errors': list, 'valid': bool}]
        """
        # Invert mapping: member_field -> csv_column
        reverse_mapping = {}
        for csv_col, member_field in field_mapping.items():
            if member_field:
                reverse_mapping[member_field] = csv_col

        results = []
        for i, row in enumerate(rows):
            mapped_data = {}
            errors = []

            for member_field, csv_col in reverse_mapping.items():
                value = row.get(csv_col, '').strip()
                mapped_data[member_field] = value

                # Validate
                validator = cls.FIELD_VALIDATORS.get(member_field)
                if validator and not validator(value):
                    errors.append(
                        _('Ligne %(row)d: valeur invalide pour %(field)s: "%(value)s"') % {
                            'row': i + 2,  # +2 for 1-indexed + header row
                            'field': member_field,
                            'value': value,
                        }
                    )

            # Required fields check
            if not mapped_data.get('first_name'):
                errors.append(_('Ligne %(row)d: prénom manquant') % {'row': i + 2})
            if not mapped_data.get('last_name'):
                errors.append(_('Ligne %(row)d: nom manquant') % {'row': i + 2})

            results.append({
                'row_num': i + 2,
                'data': mapped_data,
                'errors': errors,
                'valid': len(errors) == 0,
            })

        return results

    @classmethod
    def execute_import(cls, validated_rows, imported_by=None):
        """
        Create Member records from validated rows.

        Args:
            validated_rows: list from validate_preview (only valid rows will be imported)
            imported_by: Member who performed the import

        Returns:
            ImportHistory instance
        """
        from .models import Member, ImportHistory
        from apps.members.models import DirectoryPrivacy

        total = len(validated_rows)
        success_count = 0
        error_count = 0
        errors_json = []

        for i, row_info in enumerate(validated_rows):
            if not row_info['valid']:
                error_count += 1
                errors_json.append({
                    'row': row_info.get('row_num', i + 1),
                    'errors': [str(e) for e in row_info.get('errors', [])],
                })
                continue

            data = row_info['data']

            # Parse birth_date if provided
            birth_date = None
            if data.get('birth_date'):
                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y'):
                    try:
                        birth_date = datetime.strptime(data['birth_date'], fmt).date()
                        break
                    except (ValueError, TypeError):
                        continue

            try:
                member = Member(
                    first_name=data.get('first_name', ''),
                    last_name=data.get('last_name', ''),
                    email=data.get('email', ''),
                    phone=data.get('phone', ''),
                    birth_date=birth_date,
                    address=data.get('address', ''),
                    city=data.get('city', ''),
                    province=data.get('province', Province.QC) or Province.QC,
                    postal_code=data.get('postal_code', ''),
                    role=data.get('role', Roles.MEMBER) or Roles.MEMBER,
                    family_status=data.get('family_status', FamilyStatus.SINGLE) or FamilyStatus.SINGLE,
                )
                member.save()
                DirectoryPrivacy.objects.create(member=member)
                success_count += 1
            except Exception as e:
                error_count += 1
                errors_json.append({
                    'row': row_info.get('row_num', i + 1),
                    'errors': [str(e)],
                })

        history = ImportHistory.objects.create(
            imported_by=imported_by,
            filename='import',
            total_rows=total,
            success_count=success_count,
            error_count=error_count,
            errors_json=errors_json,
        )

        return history
